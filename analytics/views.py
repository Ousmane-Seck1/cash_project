from collections import defaultdict
from decimal import Decimal
from datetime import datetime
from io import BytesIO
import logging

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .serializers import (
    ChargeSerializer, CentreCoutSerializer, CleRepartitionSerializer, ProduitSerializer,
    CentreResponsabiliteSerializer, ActiviteSerializer,
    ResultatCalculSerializer, ExerciceSerializer, CompteChargeSerializer, FonctionSerializer,
    HopitalSerializer, UserProfileSerializer, HopitalRoleSerializer
)
from .calculations import MoteurCalculCAH
from .permissions import (
    IsControleurGestion, 
    IsComptableAnalytique,
    ReadOnly,
    CanManageUsers,
    IsGlobalAdmin,
    CanRunSensitiveOperations,
)
from .hospital_databases import ensure_hospital_database

from django.contrib.auth.models import User
from .models import (
    Charge, CentreCout, CleRepartition, Activite, Produit,
    ResultatCalcul, Exercice, CompteCharge,
    Fonction, CentreResponsabilite, Hopital, UserProfile, HopitalRole, Role, AuditLog,
    ROLE_DEFAULT_PERMISSIONS
)


def _get_user_hopital_id(user):
    if not user or not user.is_authenticated:
        return None
    if user.is_superuser:
        return None
    if hasattr(user, 'profile'):
        if user.profile.hopital_id:
            return user.profile.hopital_id
        if user.profile.hopital_role and user.profile.hopital_role.hopital_id:
            return user.profile.hopital_role.hopital_id
    return None


def _filter_for_user_hopital(queryset, user, field_path):
    if not user or not user.is_authenticated:
        return queryset.none()
    if user.is_superuser:
        return queryset

    hopital_id = _get_user_hopital_id(user)
    if not hopital_id:
        return queryset.none()

    return queryset.filter(**{field_path: hopital_id})


def _get_active_exercice_id_for_user(user):
    if not user or not user.is_authenticated:
        return None

    if user.is_superuser:
        exercice = Exercice.objects.filter(est_actif=True, est_clos=False).order_by('-annee').first()
        return exercice.id if exercice else None

    hopital_id = _get_user_hopital_id(user)
    if not hopital_id:
        return None

    exercice = Exercice.objects.filter(hopital_id=hopital_id, est_actif=True, est_clos=False).order_by('-annee').first()
    return exercice.id if exercice else None


def _get_export_hopital_name(user=None):
    if user and hasattr(user, 'profile') and user.profile.hopital:
        return user.profile.hopital.nom
    hopital = Hopital.objects.order_by('id').first()
    return hopital.nom if hopital else 'Hopital'


def _decorate_worksheet(ws, title, hopital_name, generated_at_text):
    ws.insert_rows(1, amount=4)
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Hopital: {hopital_name}"
    ws['A2'].font = Font(bold=True)
    ws['A3'] = f"Genere le: {generated_at_text}"

    header_row = 5
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    max_col = ws.max_column
    max_row = ws.max_row
    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max(12, max_length + 2), 45)

    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row):
        for cell in row:
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '#,##0'


def _decorate_workbook(wb, title, hopital_name):
    generated_at_text = datetime.now().strftime('%d/%m/%Y %H:%M')
    for ws in wb.worksheets:
        _decorate_worksheet(ws, title, hopital_name, generated_at_text)


def _build_pdf_response(filename, title, hopital_name, table_headers, table_rows, section_title=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{title}</b>", styles['Title']),
        Spacer(1, 8),
        Paragraph(f"Hopital: {hopital_name}", styles['Normal']),
        Paragraph(f"Genere le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
        Spacer(1, 10),
    ]

    if section_title:
        elements.append(Paragraph(f"<b>{section_title}</b>", styles['Heading3']))
        elements.append(Spacer(1, 6))

    table_data = [table_headers] + table_rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


def _log_audit_event(user, action, model_name, object_id, description):
    if not user or not user.is_authenticated:
        return

    allowed_actions = {choice[0] for choice in AuditLog.ACTION_CHOICES}
    if action not in allowed_actions:
        action = 'UPDATE'

    AuditLog.objects.create(
        user=user,
        action=action,
        model_name=model_name,
        object_id=int(object_id) if object_id else 0,
        description=description[:1000],
    )
    logger = logging.getLogger('analytics')
    logger.info(
        'audit_event user=%s action=%s model=%s object_id=%s desc="%s"',
        getattr(user, 'username', 'unknown'),
        action,
        model_name,
        int(object_id) if object_id else 0,
        description[:220],
    )

class ExerciceViewSet(viewsets.ModelViewSet):
    """Seul le contrôleur de gestion peut créer/modifier/supprimer des exercices"""
    queryset = Exercice.objects.all()
    serializer_class = ExerciceSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsControleurGestion()]

    def get_queryset(self):
        return _filter_for_user_hopital(Exercice.objects.all(), self.request.user, 'hopital_id').order_by('-annee')

    def perform_create(self, serializer):
        """Créer un nouvel exercice"""
        exercice = serializer.save()
        if exercice.est_actif:
            # S'assurer qu'il y a un seul exercice actif par hôpital
            Exercice.objects.filter(hopital=exercice.hopital, est_actif=True).exclude(pk=exercice.pk).update(est_actif=False)
        # Un exercice actif ne peut pas être clôturé
        if exercice.est_actif and exercice.est_clos:
            exercice.est_clos = False
            exercice.save(update_fields=['est_clos'])

    def perform_update(self, serializer):
        """Modifier un exercice (seulement s'il n'est pas clôturé)"""
        # Vérifier que l'exercice existant n'est pas clôturé
        if self.get_object().est_clos:
            raise ValidationError(
                "Cet exercice est clôturé et ne peut pas être modifié."
            )
        
        exercice = serializer.save()
        
        # Gérer l'exclusivité de l'exercice actif
        if exercice.est_actif:
            Exercice.objects.filter(hopital=exercice.hopital, est_actif=True).exclude(pk=exercice.pk).update(est_actif=False)
        
        # Un exercice actif ne peut pas être clôturé
        if exercice.est_actif and exercice.est_clos:
            exercice.est_clos = False
            exercice.save(update_fields=['est_clos'])
        
        # Si on clôture un exercice actif, le rendre inactif
        if exercice.est_clos and exercice.est_actif:
            exercice.est_actif = False
            exercice.save(update_fields=['est_actif'])

    def perform_destroy(self, instance):
        """Supprimer un exercice (impossible s'il est clôturé ou actif)"""
        if instance.est_clos:
            raise ValidationError(
                "Impossible de supprimer un exercice clôturé."
            )
        if instance.est_actif:
            raise ValidationError(
                "Impossible de supprimer l'exercice actif. Changez d'exercice actif d'abord."
            )
        instance.delete()

    @action(detail=False, methods=['post'])
    def cloturer_ouvrir(self, request):
        """Clôture l'exercice actif et ouvre le suivant pour un hôpital."""
        user = request.user
        requested_hopital_id = request.data.get('hopital')

        if user.is_superuser:
            hopital_id = requested_hopital_id
            if not hopital_id:
                return Response({'error': 'hopital requis pour un superuser'}, status=400)
        else:
            hopital_id = _get_user_hopital_id(user)
            if not hopital_id:
                return Response({'error': 'Aucun hôpital associé à votre compte.'}, status=400)

        hopital = Hopital.objects.filter(id=hopital_id).first()
        if not hopital:
            return Response({'error': 'Hôpital introuvable.'}, status=404)

        exercice_actif = Exercice.objects.filter(hopital=hopital, est_actif=True, est_clos=False).order_by('-annee').first()
        default_annee = (exercice_actif.annee + 1) if exercice_actif else datetime.now().year
        nouvelle_annee = int(request.data.get('annee') or default_annee)
        date_debut = request.data.get('date_debut') or f'{nouvelle_annee}-01-01'
        date_fin = request.data.get('date_fin') or f'{nouvelle_annee}-12-31'

        with transaction.atomic():
            if exercice_actif:
                exercice_actif.est_actif = False
                exercice_actif.est_clos = True
                exercice_actif.save(update_fields=['est_actif', 'est_clos'])

            Exercice.objects.filter(hopital=hopital, est_actif=True).update(est_actif=False)

            nouvel_exercice, created = Exercice.objects.get_or_create(
                hopital=hopital,
                annee=nouvelle_annee,
                defaults={
                    'date_debut': date_debut,
                    'date_fin': date_fin,
                    'est_actif': True,
                    'est_clos': False,
                }
            )

            if not created:
                nouvel_exercice.date_debut = date_debut
                nouvel_exercice.date_fin = date_fin
                nouvel_exercice.est_actif = True
                nouvel_exercice.est_clos = False
                nouvel_exercice.save(update_fields=['date_debut', 'date_fin', 'est_actif', 'est_clos'])

        return Response({
            'success': True,
            'hopital': {'id': hopital.id, 'nom': hopital.nom},
            'exercice_cloture': exercice_actif.annee if exercice_actif else None,
            'nouvel_exercice': {
                'id': nouvel_exercice.id,
                'annee': nouvel_exercice.annee,
                'date_debut': str(nouvel_exercice.date_debut),
                'date_fin': str(nouvel_exercice.date_fin),
            }
        })

    @action(detail=False, methods=['get'])
    def workflow_annuel_precheck(self, request):
        """Checklist prealable pour demarrage/cloture annuelle."""
        if not CanRunSensitiveOperations().has_permission(request, self):
            return Response({'error': 'Acces reserve aux profils autorises.'}, status=403)

        hopital_id = request.query_params.get('hopital')
        if request.user.is_superuser:
            if not hopital_id:
                return Response({'error': 'hopital requis pour un superuser'}, status=400)
            hopital = Hopital.objects.filter(id=hopital_id).first()
        else:
            user_hopital_id = _get_user_hopital_id(request.user)
            hopital = Hopital.objects.filter(id=user_hopital_id).first() if user_hopital_id else None

        if not hopital:
            return Response({'error': 'Hôpital introuvable.'}, status=404)

        has_exercice_actif = Exercice.objects.filter(hopital=hopital, est_actif=True, est_clos=False).exists()
        has_fonctions = Fonction.objects.filter(hopital=hopital).exists()
        has_comptes = CompteCharge.objects.filter(hopital=hopital).exists()
        centres_count = CentreCout.objects.filter(centre_responsabilite__fonction__hopital=hopital).count()
        has_ref_same_level = Hopital.objects.filter(niveau=hopital.niveau, est_reference_niveau=True).exclude(id=hopital.id).exists()

        checks = [
            {'code': 'exercice_actif', 'ok': has_exercice_actif, 'message': 'Exercice actif non cloture present.'},
            {'code': 'fonctions', 'ok': has_fonctions, 'message': 'Fonctions configurees.'},
            {'code': 'comptes', 'ok': has_comptes, 'message': 'Comptes de charges configures.'},
            {'code': 'centres', 'ok': centres_count > 0, 'message': 'Centres de cout configures.'},
            {'code': 'reference_niveau', 'ok': has_ref_same_level or hopital.est_reference_niveau, 'message': 'Reference de niveau disponible.'},
        ]

        return Response({
            'hopital': {'id': hopital.id, 'nom': hopital.nom, 'niveau': hopital.niveau},
            'ready': all(item['ok'] for item in checks[:4]),
            'checks': checks,
            'next_steps': [
                'Lancer copier_referentiel_niveau en mode preview_only pour verifier les impacts.',
                'Demarrer cloturer_ouvrir pour ouvrir le nouvel exercice.',
            ],
        }, status=200)


class CentreCoutViewSet(viewsets.ModelViewSet):
    """Gestion des centres de coûts - Permissions par rôle"""
    queryset = CentreCout.objects.all()
    serializer_class = CentreCoutSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['type_centre', 'centre_responsabilite__fonction']
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsGlobalAdmin()]

    def get_queryset(self):
        return CentreCout.objects.select_related('centre_responsabilite__fonction', 'centre_responsabilite__fonction__hopital')

    @transaction.atomic
    def perform_create(self, serializer):
        centre = serializer.save()
        parent = centre.centre_responsabilite
        if not parent:
            return

        parent_code = parent.code
        fonction_code = parent.fonction.code

        for hopital in Hopital.objects.all():
            fonction = Fonction.objects.filter(hopital=hopital, code=fonction_code).first()
            if not fonction:
                continue
            centre_responsabilite = CentreResponsabilite.objects.filter(fonction=fonction, code=parent_code).first()
            if not centre_responsabilite:
                continue
            CentreCout.objects.update_or_create(
                centre_responsabilite=centre_responsabilite,
                code=centre.code,
                defaults={
                    'libelle': centre.libelle,
                    'type_centre': centre.type_centre,
                    'unite_oeuvre': centre.unite_oeuvre,
                    'tarif': centre.tarif,
                    'ordre_cascade': centre.ordre_cascade,
                    'est_actif': centre.est_actif,
                }
            )
        _log_audit_event(
            self.request.user,
            'CREATE',
            'CentreCout',
            centre.id,
            f"Creation centre {centre.code} et synchronisation reseau.",
        )

    @transaction.atomic
    def perform_update(self, serializer):
        instance = self.get_object()
        old_code = instance.code
        old_parent_id = instance.centre_responsabilite_id

        centre = serializer.save()
        parent = centre.centre_responsabilite
        if not parent:
            return

        if centre.code != old_code or centre.centre_responsabilite_id != old_parent_id:
            raise ValidationError("Le code et le centre de responsabilite ne peuvent pas etre modifies. Supprimez puis recreez l'element.")

        new_parent_code = parent.code
        new_fonction_code = parent.fonction.code

        for hopital in Hopital.objects.all():
            new_parent = CentreResponsabilite.objects.filter(
                fonction__hopital=hopital,
                fonction__code=new_fonction_code,
                code=new_parent_code
            ).first()
            if not new_parent:
                continue

            CentreCout.objects.filter(
                centre_responsabilite=new_parent,
                code=centre.code,
            ).update(
                libelle=centre.libelle,
            )
        _log_audit_event(
            self.request.user,
            'UPDATE',
            'CentreCout',
            centre.id,
            f"Mise a jour centre {centre.code} (libelle global synchronise).",
        )

    @transaction.atomic
    def perform_destroy(self, instance):
        parent = instance.centre_responsabilite
        if not parent:
            return super().perform_destroy(instance)

        centre_code = instance.code
        parent_code = parent.code
        fonction_code = parent.fonction.code

        CentreCout.objects.filter(
            centre_responsabilite__code=parent_code,
            centre_responsabilite__fonction__code=fonction_code,
            code=centre_code
        ).delete()
        _log_audit_event(
            self.request.user,
            'DELETE',
            'CentreCout',
            instance.id,
            f"Suppression centre {centre_code} sur tout le reseau.",
        )


class CompteChargeViewSet(viewsets.ModelViewSet):
    """Gestion des comptes de charges - Permissions par rôle"""
    queryset = CompteCharge.objects.all()
    serializer_class = CompteChargeSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsGlobalAdmin()]

    def get_queryset(self):
        return CompteCharge.objects.select_related('hopital').all()

    @transaction.atomic
    def perform_create(self, serializer):
        compte = serializer.save()
        for hopital in Hopital.objects.exclude(pk=compte.hopital_id):
            CompteCharge.objects.update_or_create(
                hopital=hopital,
                numero=compte.numero,
                defaults={'libelle': compte.libelle}
            )
        _log_audit_event(
            self.request.user,
            'CREATE',
            'CompteCharge',
            compte.id,
            f"Creation compte {compte.numero} et synchronisation reseau.",
        )

    @transaction.atomic
    def perform_update(self, serializer):
        instance = self.get_object()
        old_numero = instance.numero
        compte = serializer.save()

        if compte.numero != old_numero:
            raise ValidationError("Le numero de compte ne peut pas etre modifie. Supprimez puis recreez le compte.")

        CompteCharge.objects.filter(numero=old_numero).update(libelle=compte.libelle)
        _log_audit_event(
            self.request.user,
            'UPDATE',
            'CompteCharge',
            compte.id,
            f"Mise a jour libelle compte {compte.numero} sur tout le reseau.",
        )

    @transaction.atomic
    def perform_destroy(self, instance):
        CompteCharge.objects.filter(numero=instance.numero).delete()
        _log_audit_event(
            self.request.user,
            'DELETE',
            'CompteCharge',
            instance.id,
            f"Suppression compte {instance.numero} sur tout le reseau.",
        )


class ChargeViewSet(viewsets.ModelViewSet):
    """
    Gestion des charges :
    - Contrôleur : CRUD complet
    - Comptable : CRUD sur les charges (saisie)
    - Directeur/Responsable financier : Lecture seule
    """
    queryset = Charge.objects.all()
    serializer_class = ChargeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['exercice', 'centre_cout', 'compte', 'date']
    
    def get_permissions(self):
        user = self.request.user

        if not user.is_authenticated:
            return [permissions.IsAuthenticated()]

        # Vérifier si l'utilisateur a un profil
        if not hasattr(user, 'profile'):
            # TEMPORAIRE : Créer automatiquement un profil controleur pour les utilisateurs sans profil
            from analytics.models import UserProfile
            UserProfile.objects.get_or_create(user=user, defaults={'role': 'controleur'})
            return [permissions.IsAuthenticated()]

        role = user.profile.role

        if role == 'controleur':
            return [permissions.IsAuthenticated()]  # Tout permettre
        elif role == 'comptable':
            return [IsComptableAnalytique()]  # Saisie charges uniquement
        else:
            return [ReadOnly()]  # Lecture seule pour directeur/responsable
    
    def get_queryset(self):
        """Filtrer par exercice actif par défaut"""
        queryset = _filter_for_user_hopital(
            Charge.objects.select_related('exercice', 'centre_cout', 'compte'),
            self.request.user,
            'exercice__hopital_id'
        )
        exercice_id = self.request.query_params.get('exercice')
        if not exercice_id:
            exercice_id = _get_active_exercice_id_for_user(self.request.user)
        if exercice_id:
            queryset = queryset.filter(exercice_id=exercice_id)
        return queryset.order_by('-date')
    
    def perform_create(self, serializer):
        """Assigner l'utilisateur connecté"""
        exercice = serializer.validated_data.get('exercice')
        user_hopital_id = _get_user_hopital_id(self.request.user)
        if user_hopital_id and exercice and exercice.hopital_id != user_hopital_id:
            raise PermissionDenied("Exercice invalide pour votre hôpital.")
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['get'])
    def total_par_centre(self, request):
        """Retourne le total des charges par centre"""
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'Paramètre exercice requis'}, status=400)
        
        # Vérifier les permissions de lecture
        if not request.user.is_authenticated:
            return Response({'error': 'Authentification requise'}, status=403)
        
        data = Charge.objects.filter(exercice_id=exercice_id).values(
            'centre_cout__code', 'centre_cout__libelle'
        ).annotate(total=Sum('montant'))
        
        return Response(data)


class CleRepartitionViewSet(viewsets.ModelViewSet):
    """Gestion des clés de répartition - Contrôleur uniquement pour modifications"""
    queryset = CleRepartition.objects.all()
    serializer_class = CleRepartitionSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsControleurGestion()]
    
    def get_queryset(self):
        queryset = _filter_for_user_hopital(
            CleRepartition.objects.select_related('exercice', 'centre_source', 'centre_destination'),
            self.request.user,
            'exercice__hopital_id'
        )
        exercice_id = self.request.query_params.get('exercice')
        if not exercice_id:
            exercice_id = _get_active_exercice_id_for_user(self.request.user)
        centre_source_id = self.request.query_params.get('centre_source')
        if exercice_id:
            queryset = queryset.filter(exercice_id=exercice_id)
        if centre_source_id:
            queryset = queryset.filter(centre_source_id=centre_source_id)
        return queryset

    def perform_create(self, serializer):
        exercice = serializer.validated_data.get('exercice')
        user_hopital_id = _get_user_hopital_id(self.request.user)
        if user_hopital_id and exercice and exercice.hopital_id != user_hopital_id:
            raise PermissionDenied("Exercice invalide pour votre hôpital.")
        serializer.save()
    
    @action(detail=False, methods=['post'])
    def verifier_total(self, request):
        """Vérifie que les clés d'un centre font bien 100%"""
        # Vérifier les permissions
        if not request.user.is_authenticated:
            return Response({'error': 'Authentification requise'}, status=403)
        
        exercice_id = request.data.get('exercice')
        centre_source_id = request.data.get('centre_source')
        
        if not exercice_id or not centre_source_id:
            return Response({
                'error': 'exercice et centre_source requis'
            }, status=400)
        
        total = CleRepartition.objects.filter(
            exercice_id=exercice_id,
            centre_source_id=centre_source_id
        ).aggregate(total=Sum('pourcentage'))['total'] or 0
        
        return Response({
            'centre_source': centre_source_id,
            'total_pourcentage': total,
            'est_valide': total == 100
        })


class CalculViewSet(viewsets.ViewSet):
    """
    Vues pour les calculs analytiques :
    - Contrôleur : lancer calculs + voir résultats
    - Directeur/Responsable financier : voir résultats et dashboard
    - Comptable : voir résultats
    """
    
    def get_permissions(self):
        # Par défaut, tous les utilisateurs authentifiés peuvent accéder en lecture
        if self.action in [
            'resultats',
            'tableau_analyse',
            'dashboard',
            'export_responsabilites',
            'evolution_centre',
            'export_tableau_resultats',
            'export_comparaison_tarif_cru',
            'export_charges_detaillees',
            'export_dashboard',
            'export_evolution_centre',
            'export_tableau_resultats_pdf',
            'export_comparaison_tarif_cru_pdf',
            'export_charges_detaillees_pdf',
            'export_dashboard_pdf',
            'export_evolution_centre_pdf',
            'export_responsabilites_pdf',
        ]:
            return [permissions.IsAuthenticated()]
        # Seul le contrôleur peut lancer des calculs
        return [IsControleurGestion()]

    def _resultats_queryset_for_user(self, request):
        return _filter_for_user_hopital(
            ResultatCalcul.objects.select_related('exercice', 'centre_cout', 'centre_cout__centre_responsabilite'),
            request.user,
            'exercice__hopital_id'
        )

    def _charges_queryset_for_user(self, request):
        return _filter_for_user_hopital(
            Charge.objects.select_related('exercice', 'centre_cout', 'centre_cout__centre_responsabilite', 'compte'),
            request.user,
            'exercice__hopital_id'
        )

    def _resolve_evolution_scope(self, request):
        centre_cout_id = request.query_params.get('centre_cout')
        centre_code = (request.query_params.get('centre_code') or '').strip()
        hopital_id = request.query_params.get('hopital')

        if not centre_cout_id and not centre_code:
            raise ValidationError('centre_cout ou centre_code requis')

        centres_queryset = _filter_for_user_hopital(
            CentreCout.objects.select_related('centre_responsabilite__fonction'),
            request.user,
            'centre_responsabilite__fonction__hopital_id'
        )

        centre = None
        if centre_cout_id:
            centre = centres_queryset.filter(id=centre_cout_id).first()
            if not centre:
                raise ValidationError('Centre de cout introuvable pour votre perimetre.')
            centre_code = centre.code
            hopital_id = centre.centre_responsabilite.fonction.hopital_id if centre.centre_responsabilite else None
        else:
            if hopital_id:
                centre = centres_queryset.filter(
                    code=centre_code,
                    centre_responsabilite__fonction__hopital_id=hopital_id
                ).first()
            else:
                matches = list(centres_queryset.filter(code=centre_code)[:2])
                if len(matches) > 1:
                    raise ValidationError('centre_code ambigu. Ajoutez le parametre hopital.')
                centre = matches[0] if matches else None

            if not centre:
                raise ValidationError('Centre de cout introuvable pour votre perimetre.')
            hopital_id = centre.centre_responsabilite.fonction.hopital_id if centre.centre_responsabilite else None

        if not hopital_id:
            raise ValidationError('Impossible de determiner l hopital du centre selectionne.')

        return {
            'centre_code': centre_code,
            'hopital_id': hopital_id,
            'centre_id': centre.id,
        }
    
    @action(detail=False, methods=['post'])
    def lancer_calcul(self, request):
        """Lance le calcul complet de la CAH - Contrôleur uniquement"""
        exercice_id = request.data.get('exercice')
        
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)
        
        try:
            moteur = MoteurCalculCAH(exercice_id)
            resultats = moteur.calculer_tout()
            
            return Response({
                'success': True,
                'message': 'Calcul terminé avec succès',
                'nb_centres': len(resultats)
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=500)
    
    @action(detail=False, methods=['get'])
    def resultats(self, request):
        """Retourne les résultats calculés - Tous les rôles"""
        exercice_id = request.query_params.get('exercice')
        type_centre = request.query_params.get('type_centre', None)
        
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)
        
        queryset = self._resultats_queryset_for_user(request).filter(exercice_id=exercice_id)
        
        if type_centre:
            queryset = queryset.filter(centre_cout__type_centre=type_centre)
        
        serializer = ResultatCalculSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def tableau_analyse(self, request):
        """Retourne le tableau d'analyse complet - Tous les rôles"""
        exercice_id = request.query_params.get('exercice')
        
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)
        
        try:
            moteur = MoteurCalculCAH(exercice_id)
            df = moteur.get_tableau_resultats()
            return Response(df.to_dict('records'))
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Données pour le tableau de bord - Directeur/Responsable Financier privilégié"""
        exercice_id = request.query_params.get('exercice')
        
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)
        
        # Vérifier le rôle pour dashboard restreint
        user = request.user
        if hasattr(user, 'profile'):
            role = user.profile.role
            # Le comptable n'a pas accès au dashboard complet
            if role == 'comptable':
                return Response({
                    'error': 'Accès réservé aux directeurs et responsables financiers'
                }, status=403)
        
        resultats = self._resultats_queryset_for_user(request).filter(exercice_id=exercice_id)
        
        # Statistiques globales
        total_charges = sum(r.charges_totales for r in resultats)
        total_produits = sum(r.produits or 0 for r in resultats)
        
        # Centres les plus déficitaires
        centres_deficitaires = resultats.filter(
            resultat_analytique__lt=0
        ).order_by('resultat_analytique')[:10]
        
        # Centres les plus rentables
        centres_rentables = resultats.filter(
            resultat_analytique__gt=0
        ).order_by('-resultat_analytique')[:10]
        
        return Response({
            'statistiques': {
                'total_charges': total_charges,
                'total_produits': total_produits,
                'resultat_global': total_produits - total_charges,
            },
            'centres_deficitaires': ResultatCalculSerializer(
                centres_deficitaires, many=True
            ).data,
            'centres_rentables': ResultatCalculSerializer(
                centres_rentables, many=True
            ).data,
        })

    @action(detail=False, methods=['get'])
    def export_responsabilites(self, request):
        """Exporte les résultats cumulés par centre de responsabilité en Excel."""
        exercice_id = request.query_params.get('exercice')

        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(
            exercice_id=exercice_id
        ).select_related('centre_cout__centre_responsabilite')

        regroupes = defaultdict(lambda: {
            'code': '-',
            'libelle': 'Sans centre de responsabilité',
            'nb_centres': 0,
            'charges_directes': Decimal('0'),
            'charges_indirectes': Decimal('0'),
            'charges_totales': Decimal('0'),
            'produits': Decimal('0'),
            'resultat': Decimal('0'),
        })

        for r in resultats:
            resp = r.centre_cout.centre_responsabilite
            key = resp.id if resp else 0
            data = regroupes[key]
            if resp:
                data['code'] = resp.code
                data['libelle'] = resp.libelle

            data['nb_centres'] += 1
            data['charges_directes'] += r.charges_directes or Decimal('0')
            data['charges_indirectes'] += r.charges_indirectes or Decimal('0')
            data['charges_totales'] += r.charges_totales or Decimal('0')
            data['produits'] += r.produits or Decimal('0')
            data['resultat'] += r.resultat_analytique or Decimal('0')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Par Responsabilite'

        ws.append([
            'Code Responsabilite',
            'Libelle Responsabilite',
            'Nb Centres',
            'Charges Directes',
            'Charges Indirectes',
            'Charges Totales',
            'Produits',
            'Resultat',
        ])

        for item in sorted(regroupes.values(), key=lambda x: x['code']):
            ws.append([
                item['code'],
                item['libelle'],
                item['nb_centres'],
                float(item['charges_directes']),
                float(item['charges_indirectes']),
                float(item['charges_totales']),
                float(item['produits']),
                float(item['resultat']),
            ])

        _decorate_workbook(wb, 'Resultats par Centre de Responsabilite', _get_export_hopital_name(request.user))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="resultats_par_responsabilite_exercice_{exercice_id}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def evolution_centre(self, request):
        """Retourne l'évolution des charges et produits d'un centre de coût tarifaire sur plusieurs exercices"""
        try:
            scope = self._resolve_evolution_scope(request)
        except ValidationError as exc:
            detail = exc.detail
            if isinstance(detail, (list, tuple)):
                detail = detail[0]
            return Response({'error': detail}, status=400)

        # Consolidation multi-annees pour un meme hopital via le code centre.
        resultats = self._resultats_queryset_for_user(request).filter(
            centre_cout__code=scope['centre_code'],
            exercice__hopital_id=scope['hopital_id']
        ).select_related('exercice', 'centre_cout')

        # Grouper par exercice et sommer les charges/produits
        evolution = defaultdict(lambda: {
            'exercice': None,
            'annee': None,
            'charges_totales': Decimal('0'),
            'produits': Decimal('0'),
        })

        for resultat in resultats.order_by('exercice__annee'):
            ex_id = resultat.exercice.id
            evolution[ex_id]['exercice'] = resultat.exercice.id
            evolution[ex_id]['annee'] = resultat.exercice.annee
            evolution[ex_id]['charges_totales'] += resultat.charges_totales or Decimal('0')
            evolution[ex_id]['produits'] += resultat.produits or Decimal('0')

        # Convertir en liste triée par année
        data = sorted(
            [{'exercice': v['exercice'], 'annee': v['annee'], 'charges_totales': float(v['charges_totales']), 'produits': float(v['produits'])} for v in evolution.values()],
            key=lambda x: x['annee']
        )

        return Response({
            'centre_cout': scope['centre_id'],
            'centre_code': scope['centre_code'],
            'hopital': scope['hopital_id'],
            'evolution': data,
        })

    @action(detail=False, methods=['get'])
    def export_tableau_resultats(self, request):
        """Exporte le tableau des résultats par centre de coût."""
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(exercice_id=exercice_id).select_related('centre_cout')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Tableau Resultats'

        ws.append([
            'Code Centre', 'Libelle Centre', 'Type Centre',
            'Charges Directes', 'Charges Indirectes', 'Charges Totales',
            'Produits', 'Resultat Analytique'
        ])

        total_directes = Decimal('0')
        total_indirectes = Decimal('0')
        total_charges = Decimal('0')
        total_produits = Decimal('0')
        total_resultat = Decimal('0')

        for r in sorted(resultats, key=lambda item: item.centre_cout.code or ''):
            charges_directes = r.charges_directes or Decimal('0')
            charges_indirectes = r.charges_indirectes or Decimal('0')
            charges_totales = r.charges_totales or Decimal('0')
            produits = r.produits or Decimal('0')
            resultat = r.resultat_analytique or Decimal('0')

            total_directes += charges_directes
            total_indirectes += charges_indirectes
            total_charges += charges_totales
            total_produits += produits
            total_resultat += resultat

            ws.append([
                r.centre_cout.code,
                r.centre_cout.libelle,
                r.centre_cout.type_centre,
                float(charges_directes),
                float(charges_indirectes),
                float(charges_totales),
                float(produits),
                float(resultat),
            ])

        ws.append([
            'TOTAL', '', '',
            float(total_directes), float(total_indirectes), float(total_charges),
            float(total_produits), float(total_resultat)
        ])

        _decorate_workbook(wb, 'Tableau des Resultats', _get_export_hopital_name(request.user))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tableau_resultats_exercice_{exercice_id}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_comparaison_tarif_cru(self, request):
        """Exporte la comparaison tarif vs CRU pour les centres tarifaires."""
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(
            exercice_id=exercice_id,
            centre_cout__type_centre__in=['CT_MT', 'CT_CL']
        ).select_related('centre_cout')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Comparaison Tarif CRU'

        ws.append([
            'Centre de Cout', 'Total Charges', 'Volume Activites', 'Unite d Oeuvre',
            'CRU', 'Tarif', 'Ecart (Tarif-CRU)'
        ])

        total_charges = Decimal('0')

        for r in sorted(resultats, key=lambda item: item.centre_cout.code or ''):
            charges_totales = r.charges_totales or Decimal('0')
            volume = r.volume_activite
            cru = r.cout_revient_unitaire
            tarif = r.centre_cout.tarif
            ecart = None
            if cru is not None and tarif is not None:
                ecart = tarif - cru

            total_charges += charges_totales

            ws.append([
                f"{r.centre_cout.code} - {r.centre_cout.libelle}",
                float(charges_totales),
                float(volume) if volume is not None else None,
                r.centre_cout.unite_oeuvre,
                float(cru) if cru is not None else None,
                float(tarif) if tarif is not None else None,
                float(ecart) if ecart is not None else None,
            ])

        ws.append(['TOTAL', float(total_charges), None, None, None, None, None])

        _decorate_workbook(wb, 'Comparaison Tarif vs CRU', _get_export_hopital_name(request.user))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="comparaison_tarif_cru_exercice_{exercice_id}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_charges_detaillees(self, request):
        """Exporte les charges détaillées avec filtres optionnels par responsabilité et centre."""
        exercice_id = request.query_params.get('exercice')
        resp_id = request.query_params.get('centre_responsabilite')
        centre_id = request.query_params.get('centre_cout')

        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        charges = self._charges_queryset_for_user(request).filter(exercice_id=exercice_id).select_related(
            'centre_cout__centre_responsabilite',
            'compte'
        )

        if resp_id:
            charges = charges.filter(centre_cout__centre_responsabilite_id=resp_id)
        if centre_id:
            charges = charges.filter(centre_cout_id=centre_id)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Charges Detaillees'

        ws.append([
            'Date', 'Code Resp', 'Libelle Resp', 'Code Centre', 'Libelle Centre',
            'Compte', 'Designation', 'Montant'
        ])

        total_montant = Decimal('0')

        for c in charges.order_by('date', 'centre_cout__code'):
            montant = c.montant or Decimal('0')
            total_montant += montant
            resp = c.centre_cout.centre_responsabilite if c.centre_cout else None

            ws.append([
                c.date.isoformat() if c.date else '',
                resp.code if resp else '',
                resp.libelle if resp else '',
                c.centre_cout.code if c.centre_cout else '',
                c.centre_cout.libelle if c.centre_cout else '',
                f"{c.compte.numero} - {c.compte.libelle}" if c.compte else '',
                c.designation,
                float(montant),
            ])

        ws.append(['TOTAL', '', '', '', '', '', '', float(total_montant)])

        _decorate_workbook(wb, 'Charges Detaillees', _get_export_hopital_name(request.user))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="charges_detaillees_exercice_{exercice_id}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_dashboard(self, request):
        """Exporte les indicateurs dashboard et top centres."""
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(exercice_id=exercice_id).select_related('centre_cout')

        total_charges = sum((r.charges_totales or Decimal('0')) for r in resultats)
        total_produits = sum((r.produits or Decimal('0')) for r in resultats)
        resultat_global = total_produits - total_charges

        centres_deficitaires = resultats.filter(resultat_analytique__lt=0).order_by('resultat_analytique')[:10]
        centres_rentables = resultats.filter(resultat_analytique__gt=0).order_by('-resultat_analytique')[:10]

        wb = Workbook()
        ws_resume = wb.active
        ws_resume.title = 'Dashboard Resume'
        ws_resume.append(['Indicateur', 'Valeur'])
        ws_resume.append(['Total Charges', float(total_charges)])
        ws_resume.append(['Total Produits', float(total_produits)])
        ws_resume.append(['Resultat Global', float(resultat_global)])

        ws_excedent = wb.create_sheet('Top Excedentaires')
        ws_excedent.append(['Rang', 'Code Centre', 'Libelle Centre', 'Resultat'])
        for idx, c in enumerate(centres_rentables, start=1):
            ws_excedent.append([
                idx,
                c.centre_cout.code if c.centre_cout else '',
                c.centre_cout.libelle if c.centre_cout else '',
                float(c.resultat_analytique or 0),
            ])

        ws_deficit = wb.create_sheet('Top Deficitaires')
        ws_deficit.append(['Rang', 'Code Centre', 'Libelle Centre', 'Resultat'])
        for idx, c in enumerate(centres_deficitaires, start=1):
            ws_deficit.append([
                idx,
                c.centre_cout.code if c.centre_cout else '',
                c.centre_cout.libelle if c.centre_cout else '',
                float(c.resultat_analytique or 0),
            ])

        _decorate_workbook(wb, 'Dashboard Analytique', _get_export_hopital_name(request.user))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="dashboard_exercice_{exercice_id}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_evolution_centre(self, request):
        """Exporte l'évolution d'un centre tarifaire sur tous les exercices."""
        try:
            scope = self._resolve_evolution_scope(request)
        except ValidationError as exc:
            detail = exc.detail
            if isinstance(detail, (list, tuple)):
                detail = detail[0]
            return Response({'error': detail}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(
            centre_cout__code=scope['centre_code'],
            exercice__hopital_id=scope['hopital_id']
        ).select_related('exercice', 'centre_cout').order_by('exercice__annee')

        evolution = defaultdict(lambda: {
            'annee': None,
            'charges_totales': Decimal('0'),
            'produits': Decimal('0'),
        })

        centre_label = ''
        for r in resultats:
            if not centre_label and r.centre_cout:
                centre_label = f"{r.centre_cout.code}_{r.centre_cout.libelle}".replace(' ', '_')

            key = r.exercice.id
            evolution[key]['annee'] = r.exercice.annee
            evolution[key]['charges_totales'] += r.charges_totales or Decimal('0')
            evolution[key]['produits'] += r.produits or Decimal('0')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Evolution'
        ws.append(['Annee', 'Charges Totales', 'Produits'])

        for item in sorted(evolution.values(), key=lambda x: x['annee'] or 0):
            ws.append([
                item['annee'],
                float(item['charges_totales']),
                float(item['produits']),
            ])

        _decorate_workbook(wb, 'Evolution des Centres Tarifaires', _get_export_hopital_name(request.user))

        file_suffix = centre_label or f"centre_{scope['centre_id']}"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="evolution_{file_suffix}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_tableau_resultats_pdf(self, request):
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(exercice_id=exercice_id).select_related('centre_cout')
        rows = []
        for r in sorted(resultats, key=lambda item: item.centre_cout.code or ''):
            rows.append([
                r.centre_cout.code,
                r.centre_cout.libelle,
                r.centre_cout.type_centre,
                f"{float(r.charges_directes or 0):,.0f}",
                f"{float(r.charges_indirectes or 0):,.0f}",
                f"{float(r.charges_totales or 0):,.0f}",
                f"{float(r.produits or 0):,.0f}",
                f"{float(r.resultat_analytique or 0):,.0f}",
            ])

        return _build_pdf_response(
            f'tableau_resultats_exercice_{exercice_id}.pdf',
            'Tableau des Resultats',
            _get_export_hopital_name(request.user),
            ['Code', 'Libelle', 'Type', 'Ch. Directes', 'Ch. Indirectes', 'Ch. Totales', 'Produits', 'Resultat'],
            rows or [['-', '-', '-', '-', '-', '-', '-', '-']]
        )

    @action(detail=False, methods=['get'])
    def export_comparaison_tarif_cru_pdf(self, request):
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(
            exercice_id=exercice_id,
            centre_cout__type_centre__in=['CT_MT', 'CT_CL']
        ).select_related('centre_cout')

        rows = []
        for r in sorted(resultats, key=lambda item: item.centre_cout.code or ''):
            cru = r.cout_revient_unitaire
            tarif = r.centre_cout.tarif
            ecart = (tarif - cru) if cru is not None and tarif is not None else None
            rows.append([
                f"{r.centre_cout.code} - {r.centre_cout.libelle}",
                f"{float(r.charges_totales or 0):,.0f}",
                f"{float(r.volume_activite):,.0f}" if r.volume_activite is not None else '-',
                r.centre_cout.unite_oeuvre or '-',
                f"{float(cru):,.0f}" if cru is not None else '-',
                f"{float(tarif):,.0f}" if tarif is not None else '-',
                f"{float(ecart):,.0f}" if ecart is not None else '-',
            ])

        return _build_pdf_response(
            f'comparaison_tarif_cru_exercice_{exercice_id}.pdf',
            'Comparaison Tarif / CRU',
            _get_export_hopital_name(request.user),
            ['Centre', 'Total Charges', 'Volume', 'UO', 'CRU', 'Tarif', 'Ecart'],
            rows or [['-', '-', '-', '-', '-', '-', '-']]
        )

    @action(detail=False, methods=['get'])
    def export_charges_detaillees_pdf(self, request):
        exercice_id = request.query_params.get('exercice')
        resp_id = request.query_params.get('centre_responsabilite')
        centre_id = request.query_params.get('centre_cout')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        charges = self._charges_queryset_for_user(request).filter(exercice_id=exercice_id).select_related(
            'centre_cout__centre_responsabilite', 'compte'
        )
        if resp_id:
            charges = charges.filter(centre_cout__centre_responsabilite_id=resp_id)
        if centre_id:
            charges = charges.filter(centre_cout_id=centre_id)

        rows = []
        for c in charges.order_by('date', 'centre_cout__code'):
            resp = c.centre_cout.centre_responsabilite if c.centre_cout else None
            rows.append([
                c.date.isoformat() if c.date else '-',
                f"{resp.code if resp else '-'} - {resp.libelle if resp else '-'}",
                f"{c.centre_cout.code if c.centre_cout else '-'} - {c.centre_cout.libelle if c.centre_cout else '-'}",
                f"{c.compte.numero if c.compte else '-'}",
                c.designation or '-',
                f"{float(c.montant or 0):,.0f}",
            ])

        return _build_pdf_response(
            f'charges_detaillees_exercice_{exercice_id}.pdf',
            'Charges Detaillees',
            _get_export_hopital_name(request.user),
            ['Date', 'Responsabilite', 'Centre', 'Compte', 'Designation', 'Montant'],
            rows or [['-', '-', '-', '-', '-', '-']]
        )

    @action(detail=False, methods=['get'])
    def export_dashboard_pdf(self, request):
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(exercice_id=exercice_id).select_related('centre_cout')
        total_charges = sum((r.charges_totales or Decimal('0')) for r in resultats)
        total_produits = sum((r.produits or Decimal('0')) for r in resultats)
        resultat_global = total_produits - total_charges
        centres_deficitaires = resultats.filter(resultat_analytique__lt=0).order_by('resultat_analytique')[:10]
        centres_rentables = resultats.filter(resultat_analytique__gt=0).order_by('-resultat_analytique')[:10]

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        hopital_name = _get_export_hopital_name(request.user)
        elements = [
            Paragraph('<b>Dashboard Analytique</b>', styles['Title']),
            Spacer(1, 8),
            Paragraph(f"Hopital: {hopital_name}", styles['Normal']),
            Paragraph(f"Genere le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 10),
            Paragraph('<b>Indicateurs</b>', styles['Heading3']),
            Spacer(1, 6),
        ]

        table_indicateurs = Table([
            ['Indicateur', 'Valeur'],
            ['Total Charges', f"{float(total_charges):,.0f}"],
            ['Total Produits', f"{float(total_produits):,.0f}"],
            ['Resultat Global', f"{float(resultat_global):,.0f}"],
        ], repeatRows=1)
        table_indicateurs.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table_indicateurs)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph('<b>Top 10 Centres Excedentaires</b>', styles['Heading3']))
        elements.append(Spacer(1, 6))
        rows_excedent = [['Rang', 'Code Centre', 'Libelle Centre', 'Resultat']] + [
            [
                idx,
                c.centre_cout.code if c.centre_cout else '-',
                c.centre_cout.libelle if c.centre_cout else '-',
                f"{float(c.resultat_analytique or 0):,.0f}",
            ]
            for idx, c in enumerate(centres_rentables, start=1)
        ]
        if len(rows_excedent) == 1:
            rows_excedent.append(['-', '-', '-', '-'])
        table_excedent = Table(rows_excedent, repeatRows=1)
        table_excedent.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table_excedent)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph('<b>Top 10 Centres Deficitaires</b>', styles['Heading3']))
        elements.append(Spacer(1, 6))
        rows_deficit = [['Rang', 'Code Centre', 'Libelle Centre', 'Resultat']] + [
            [
                idx,
                c.centre_cout.code if c.centre_cout else '-',
                c.centre_cout.libelle if c.centre_cout else '-',
                f"{float(c.resultat_analytique or 0):,.0f}",
            ]
            for idx, c in enumerate(centres_deficitaires, start=1)
        ]
        if len(rows_deficit) == 1:
            rows_deficit.append(['-', '-', '-', '-'])
        table_deficit = Table(rows_deficit, repeatRows=1)
        table_deficit.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table_deficit)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="dashboard_exercice_{exercice_id}.pdf"'
        response.write(pdf)
        return response

    @action(detail=False, methods=['get'])
    def export_responsabilites_pdf(self, request):
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response({'error': 'exercice requis'}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(
            exercice_id=exercice_id
        ).select_related('centre_cout__centre_responsabilite')

        regroupes = defaultdict(lambda: {
            'code': '-',
            'libelle': 'Sans centre de responsabilite',
            'nb_centres': 0,
            'charges_directes': Decimal('0'),
            'charges_indirectes': Decimal('0'),
            'charges_totales': Decimal('0'),
            'produits': Decimal('0'),
            'resultat': Decimal('0'),
        })

        for r in resultats:
            resp = r.centre_cout.centre_responsabilite
            key = resp.id if resp else 0
            data = regroupes[key]
            if resp:
                data['code'] = resp.code
                data['libelle'] = resp.libelle

            data['nb_centres'] += 1
            data['charges_directes'] += r.charges_directes or Decimal('0')
            data['charges_indirectes'] += r.charges_indirectes or Decimal('0')
            data['charges_totales'] += r.charges_totales or Decimal('0')
            data['produits'] += r.produits or Decimal('0')
            data['resultat'] += r.resultat_analytique or Decimal('0')

        rows = []
        for item in sorted(regroupes.values(), key=lambda x: x['code']):
            rows.append([
                item['code'],
                item['libelle'],
                item['nb_centres'],
                f"{float(item['charges_directes']):,.0f}",
                f"{float(item['charges_indirectes']):,.0f}",
                f"{float(item['charges_totales']):,.0f}",
                f"{float(item['produits']):,.0f}",
                f"{float(item['resultat']):,.0f}",
            ])

        return _build_pdf_response(
            f'resultats_par_responsabilite_exercice_{exercice_id}.pdf',
            'Resultats par Centre de Responsabilite',
            _get_export_hopital_name(request.user),
            ['Code Resp.', 'Libelle Resp.', 'Nb Centres', 'Ch. Directes', 'Ch. Indirectes', 'Ch. Totales', 'Produits', 'Resultat'],
            rows or [['-', '-', '-', '-', '-', '-', '-', '-']]
        )

    @action(detail=False, methods=['get'])
    def export_evolution_centre_pdf(self, request):
        try:
            scope = self._resolve_evolution_scope(request)
        except ValidationError as exc:
            detail = exc.detail
            if isinstance(detail, (list, tuple)):
                detail = detail[0]
            return Response({'error': detail}, status=400)

        resultats = self._resultats_queryset_for_user(request).filter(
            centre_cout__code=scope['centre_code'],
            exercice__hopital_id=scope['hopital_id']
        ).select_related('exercice', 'centre_cout').order_by('exercice__annee')

        evolution = defaultdict(lambda: {
            'annee': None,
            'charges_totales': Decimal('0'),
            'produits': Decimal('0'),
        })
        for r in resultats:
            key = r.exercice.id
            evolution[key]['annee'] = r.exercice.annee
            evolution[key]['charges_totales'] += r.charges_totales or Decimal('0')
            evolution[key]['produits'] += r.produits or Decimal('0')

        rows = []
        for item in sorted(evolution.values(), key=lambda x: x['annee'] or 0):
            rows.append([
                item['annee'],
                f"{float(item['charges_totales']):,.0f}",
                f"{float(item['produits']):,.0f}",
            ])

        return _build_pdf_response(
            f"evolution_centre_{scope['centre_code']}_hopital_{scope['hopital_id']}.pdf",
            'Evolution des Centres Tarifaires',
            _get_export_hopital_name(request.user),
            ['Annee', 'Charges Totales', 'Produits'],
            rows or [['-', '-', '-']]
        )


# ========== GESTION DES UTILISATEURS ==========

class UserViewSet(viewsets.ModelViewSet):
    """
    CRUD des utilisateurs et de leurs profils.
    Seul le contrôleur de gestion peut gérer les utilisateurs.
    """
    queryset = UserProfile.objects.select_related('user', 'hopital', 'hopital_role').all()
    serializer_class = UserProfileSerializer

    def get_permissions(self):
        return [CanManageUsers()]

    def _ensure_missing_profiles(self):
        users_without_profile = User.objects.filter(profile__isnull=True)
        for user in users_without_profile:
            default_role = Role.CONTROLEUR_GESTION if user.is_superuser else Role.AGENT_SAISIE
            UserProfile.objects.get_or_create(
                user=user,
                defaults={'role': default_role, 'hopital': None, 'hopital_role': None}
            )

    def get_queryset(self):
        self._ensure_missing_profiles()
        queryset = UserProfile.objects.select_related('user', 'hopital', 'hopital_role').all()
        user = self.request.user
        if user.is_superuser:
            return queryset
        user_hopital_id = _get_user_hopital_id(user)
        if user_hopital_id:
            return queryset.filter(hopital_id=user_hopital_id)
        return queryset.none()

    def destroy(self, request, *args, **kwargs):
        profile = self.get_object()
        if profile.user == request.user:
            return Response({'error': 'Vous ne pouvez pas supprimer votre propre compte.'}, status=400)
        user = profile.user
        profile.delete()
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class HopitalRoleViewSet(viewsets.ModelViewSet):
    """Gestion des rôles flexibles et permissions par hôpital."""
    queryset = HopitalRole.objects.select_related('hopital').all()
    serializer_class = HopitalRoleSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['hopital', 'code', 'est_actif']

    def get_permissions(self):
        return [CanManageUsers()]

    def get_queryset(self):
        queryset = HopitalRole.objects.select_related('hopital').all()
        user = self.request.user
        if user.is_superuser:
            return queryset
        user_hopital_id = _get_user_hopital_id(user)
        if user_hopital_id:
            return queryset.filter(hopital_id=user_hopital_id)
        return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.is_superuser:
            serializer.save()
            return

        user_hopital_id = _get_user_hopital_id(user)
        if not user_hopital_id:
            raise PermissionDenied("Aucun hôpital associé à votre compte.")

        serializer.save(hopital_id=user_hopital_id)

    def perform_update(self, serializer):
        user = self.request.user
        if user.is_superuser:
            serializer.save()
            return

        user_hopital_id = _get_user_hopital_id(user)
        if not user_hopital_id:
            raise PermissionDenied("Aucun hôpital associé à votre compte.")

        serializer.save(hopital_id=user_hopital_id)

    @action(detail=False, methods=['get'])
    def permission_catalog(self, request):
        return Response(HopitalRole.permission_catalog())


# ========== VUE D'ACCUEIL =========
def index_view(request):
    return render(request, 'analytics/index.html')


# ========== VIEWSETS ==========

class FonctionViewSet(viewsets.ModelViewSet):
    """Gestion des fonctions (niveaux 1 du CASH)"""
    queryset = Fonction.objects.all()
    serializer_class = FonctionSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsGlobalAdmin()]

    def get_queryset(self):
        return Fonction.objects.select_related('hopital').order_by('code', 'hopital__code')

    @transaction.atomic
    def perform_create(self, serializer):
        fonction = serializer.save()
        for hopital in Hopital.objects.exclude(pk=fonction.hopital_id):
            Fonction.objects.update_or_create(
                hopital=hopital,
                code=fonction.code,
                defaults={'libelle': fonction.libelle}
            )
        _log_audit_event(
            self.request.user,
            'CREATE',
            'Fonction',
            fonction.id,
            f"Creation fonction {fonction.code} et synchronisation reseau.",
        )

    @transaction.atomic
    def perform_update(self, serializer):
        instance = self.get_object()
        old_code = instance.code
        fonction = serializer.save()

        if fonction.code != old_code:
            raise ValidationError("Le code fonction ne peut pas etre modifie. Supprimez puis recreez la fonction.")

        Fonction.objects.filter(code=old_code).update(libelle=fonction.libelle)
        _log_audit_event(
            self.request.user,
            'UPDATE',
            'Fonction',
            fonction.id,
            f"Mise a jour libelle fonction {fonction.code} sur tout le reseau.",
        )

    @transaction.atomic
    def perform_destroy(self, instance):
        Fonction.objects.filter(code=instance.code).delete()
        _log_audit_event(
            self.request.user,
            'DELETE',
            'Fonction',
            instance.id,
            f"Suppression fonction {instance.code} sur tout le reseau.",
        )



class CentreResponsabiliteViewSet(viewsets.ModelViewSet):
    queryset = CentreResponsabilite.objects.all()
    serializer_class = CentreResponsabiliteSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [IsGlobalAdmin()]

    def get_queryset(self):
        return CentreResponsabilite.objects.select_related('fonction', 'fonction__hopital').order_by('fonction__code', 'code', 'fonction__hopital__code')

    @transaction.atomic
    def perform_create(self, serializer):
        centre = serializer.save()
        fonction_code = centre.fonction.code

        for hopital in Hopital.objects.all():
            fonction = Fonction.objects.filter(hopital=hopital, code=fonction_code).first()
            if not fonction:
                continue
            CentreResponsabilite.objects.update_or_create(
                fonction=fonction,
                code=centre.code,
                defaults={'libelle': centre.libelle}
            )
        _log_audit_event(
            self.request.user,
            'CREATE',
            'CentreResponsabilite',
            centre.id,
            f"Creation centre responsabilite {centre.code} et synchronisation reseau.",
        )

    @transaction.atomic
    def perform_update(self, serializer):
        instance = self.get_object()
        old_code = instance.code
        old_fonction_id = instance.fonction_id
        centre = serializer.save()

        if centre.code != old_code or centre.fonction_id != old_fonction_id:
            raise ValidationError("Le code et la fonction ne peuvent pas etre modifies. Supprimez puis recreez l'element.")

        CentreResponsabilite.objects.filter(
            fonction__code=centre.fonction.code,
            code=old_code
        ).update(libelle=centre.libelle)
        _log_audit_event(
            self.request.user,
            'UPDATE',
            'CentreResponsabilite',
            centre.id,
            f"Mise a jour libelle centre responsabilite {centre.code} sur tout le reseau.",
        )

    @transaction.atomic
    def perform_destroy(self, instance):
        CentreResponsabilite.objects.filter(
            fonction__code=instance.fonction.code,
            code=instance.code
        ).delete()
        _log_audit_event(
            self.request.user,
            'DELETE',
            'CentreResponsabilite',
            instance.id,
            f"Suppression centre responsabilite {instance.code} sur tout le reseau.",
        )
    

class ProduitViewSet(viewsets.ModelViewSet):
    """
    Gestion des produits (recettes) par centre tarifaire
    """
    queryset = Produit.objects.all()
    serializer_class = ProduitSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['exercice', 'centre_cout', 'periode']
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        # Contrôleur et comptable peuvent saisir
        return [permissions.IsAuthenticated()]  # Simplifié pour test
    
    def get_queryset(self):
        queryset = _filter_for_user_hopital(
            Produit.objects.select_related('exercice', 'centre_cout'),
            self.request.user,
            'exercice__hopital_id'
        )
        exercice_id = self.request.query_params.get('exercice')
        if not exercice_id:
            exercice_id = _get_active_exercice_id_for_user(self.request.user)
        if exercice_id:
            queryset = queryset.filter(exercice_id=exercice_id)
        return queryset.order_by('centre_cout__code', 'periode')
    
    def perform_create(self, serializer):
        # Assigner l'exercice actif si non fourni
        exercice_id = self.request.data.get('exercice')
        user_hopital_id = _get_user_hopital_id(self.request.user)
        if not exercice_id:
            exercice = Exercice.objects.filter(est_actif=True)
            if user_hopital_id:
                exercice = exercice.filter(hopital_id=user_hopital_id)
            exercice = exercice.first()
            serializer.save(exercice=exercice)
        else:
            exercice = serializer.validated_data.get('exercice')
            if user_hopital_id and exercice and exercice.hopital_id != user_hopital_id:
                raise PermissionDenied("Exercice invalide pour votre hôpital.")
            serializer.save()


class ActiviteViewSet(viewsets.ModelViewSet):
    """
    Gestion des volumes d'activités par centre tarifaire
    """
    queryset = Activite.objects.all()
    serializer_class = ActiviteSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['exercice', 'centre_cout', 'periode']
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]  # Simplifié pour test
    
    def get_queryset(self):
        queryset = _filter_for_user_hopital(
            Activite.objects.select_related('exercice', 'centre_cout'),
            self.request.user,
            'exercice__hopital_id'
        )
        exercice_id = self.request.query_params.get('exercice')
        if not exercice_id:
            exercice_id = _get_active_exercice_id_for_user(self.request.user)
        if exercice_id:
            queryset = queryset.filter(exercice_id=exercice_id)
        return queryset.order_by('centre_cout__code', 'periode')
    
    def perform_create(self, serializer):
        exercice_id = self.request.data.get('exercice')
        user_hopital_id = _get_user_hopital_id(self.request.user)
        if not exercice_id:
            exercice = Exercice.objects.filter(est_actif=True)
            if user_hopital_id:
                exercice = exercice.filter(hopital_id=user_hopital_id)
            exercice = exercice.first()
            serializer.save(exercice=exercice)
        else:
            exercice = serializer.validated_data.get('exercice')
            if user_hopital_id and exercice and exercice.hopital_id != user_hopital_id:
                raise PermissionDenied("Exercice invalide pour votre hôpital.")
            serializer.save()


# ========== HÔPITAL ==========

class HopitalViewSet(viewsets.ModelViewSet):
    """
    Gestion des données de l'hôpital.
    Seul le contrôleur de gestion peut modifier.
    """
    queryset = Hopital.objects.all()
    serializer_class = HopitalSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]

    def get_queryset(self):
        return _filter_for_user_hopital(Hopital.objects.all(), self.request.user, 'id')

    def create(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            raise PermissionDenied("Seul un superuser peut créer un hôpital.")

        response = super().create(request, *args, **kwargs)
        hopital_id = response.data.get('id')
        if hopital_id:
            hopital = Hopital.objects.filter(id=hopital_id).first()
            if hopital:
                response.data['database'] = {
                    'alias': hopital.database_alias,
                    'name': hopital.database_name,
                }
        return response

    def perform_create(self, serializer):
        if not self.request.user.is_superuser:
            raise PermissionDenied("Seul un superuser peut créer un hôpital.")
        data = serializer.validated_data
        self._check_reference_conflict(
            niveau=data.get('niveau', Hopital.NIVEAU_1),
            est_reference_niveau=data.get('est_reference_niveau', False),
            current_id=None,
        )
        hopital = serializer.save()
        _log_audit_event(
            self.request.user,
            'CREATE',
            'Hopital',
            hopital.id,
            f"Creation hopital {hopital.nom} ({hopital.code}) niveau={hopital.niveau} ref_niveau={hopital.est_reference_niveau}.",
        )

    def perform_update(self, serializer):
        if not self.request.user.is_superuser:
            raise PermissionDenied("Seul un superuser peut modifier un hôpital.")
        current = self.get_object()
        data = serializer.validated_data
        niveau = data.get('niveau', current.niveau)
        est_reference_niveau = data.get('est_reference_niveau', current.est_reference_niveau)
        self._check_reference_conflict(
            niveau=niveau,
            est_reference_niveau=est_reference_niveau,
            current_id=current.id,
        )
        hopital = serializer.save()
        _log_audit_event(
            self.request.user,
            'UPDATE',
            'Hopital',
            hopital.id,
            f"Mise a jour hopital {hopital.nom} ({hopital.code}) niveau={hopital.niveau} ref_niveau={hopital.est_reference_niveau}.",
        )

    def _check_reference_conflict(self, niveau, est_reference_niveau, current_id=None):
        if not est_reference_niveau:
            return
        duplicate = Hopital.objects.filter(
            niveau=niveau,
            est_reference_niveau=True,
        )
        if current_id:
            duplicate = duplicate.exclude(id=current_id)
        duplicate = duplicate.first()
        if duplicate:
            raise ValidationError(
                f"Le niveau {niveau} a deja un hopital de reference: {duplicate.nom}."
            )

    @action(detail=True, methods=['post'])
    def provision_database(self, request, pk=None):
        """(Re)provisionne la base dediee de l'hopital."""
        hopital = self.get_object()
        try:
            db_info = ensure_hospital_database(hopital)
        except Exception as exc:
            return Response({'error': f"Provisionnement impossible: {exc}"}, status=500)

        _log_audit_event(
            request.user,
            'UPDATE',
            'Hopital',
            hopital.id,
            f"Provisionnement base dediee: alias={db_info.get('alias')} name={db_info.get('name')}.",
        )

        return Response({
            'success': True,
            'hopital': {'id': hopital.id, 'nom': hopital.nom},
            'database': db_info,
        })

    @action(detail=True, methods=['post'])
    def reinitialiser_saisie(self, request, pk=None):
        """Supprime les données de saisie d'un hôpital (sans toucher aux résultats)."""
        if str(request.data.get('confirmation_code') or '').strip().upper() != 'CONFIRMER':
            return Response({'error': 'confirmation_code=CONFIRMER requis pour cette action critique.'}, status=400)

        hopital = self.get_object()
        exercices_ids = list(Exercice.objects.filter(hopital=hopital).values_list('id', flat=True))

        if not exercices_ids:
            return Response({'success': True, 'message': 'Aucun exercice trouvé pour cet hôpital.', 'details': {}}, status=200)

        deleted_charges, _ = Charge.objects.filter(exercice_id__in=exercices_ids).delete()
        deleted_cles, _ = CleRepartition.objects.filter(exercice_id__in=exercices_ids).delete()
        deleted_produits, _ = Produit.objects.filter(exercice_id__in=exercices_ids).delete()
        deleted_activites, _ = Activite.objects.filter(exercice_id__in=exercices_ids).delete()

        _log_audit_event(
            request.user,
            'DELETE',
            'Hopital',
            hopital.id,
            f"Reinitialisation saisie: charges={deleted_charges}, cles={deleted_cles}, produits={deleted_produits}, activites={deleted_activites}.",
        )

        return Response({
            'success': True,
            'message': f'Données de saisie réinitialisées pour {hopital.nom}.',
            'details': {
                'charges': deleted_charges,
                'cles_repartition': deleted_cles,
                'produits': deleted_produits,
                'activites': deleted_activites,
            }
        }, status=200)

    @action(detail=True, methods=['post'])
    def dupliquer_configuration(self, request, pk=None):
        """Duplique la configuration (fonctions, centres, comptes) depuis un hôpital source."""
        target_hopital = self.get_object()
        source_hopital_id = request.data.get('source_hopital_id')
        force = str(request.data.get('force', 'true')).strip().lower() in {'1', 'true', 'yes', 'oui', 'on'}
        if force and str(request.data.get('confirmation_code') or '').strip().upper() != 'CONFIRMER':
            return Response({'error': 'confirmation_code=CONFIRMER requis quand force=true.'}, status=400)

        if source_hopital_id:
            source_hopital = Hopital.objects.filter(id=source_hopital_id).first()
        else:
            source_hopital = Hopital.objects.exclude(id=target_hopital.id).order_by('id').first()

        if not source_hopital:
            return Response({'error': 'Aucun hôpital source disponible.'}, status=400)

        preview_only = str(request.data.get('preview_only', 'false')).strip().lower() in {'1', 'true', 'yes', 'oui', 'on'}
        diff = self._build_configuration_diff(source_hopital, target_hopital)
        if preview_only:
            return Response({
                'success': True,
                'preview_only': True,
                'source_hopital': {'id': source_hopital.id, 'nom': source_hopital.nom},
                'target_hopital': {'id': target_hopital.id, 'nom': target_hopital.nom},
                'force_center_fields': force,
                'diff': diff,
            }, status=200)

        created, updated = self._copy_configuration_between_hospitals(
            source_hopital=source_hopital,
            target_hopital=target_hopital,
            force_center_fields=force,
        )
        _log_audit_event(
            request.user,
            'UPDATE',
            'Hopital',
            target_hopital.id,
            f"Duplication configuration depuis {source_hopital.code} vers {target_hopital.code} (force={force}).",
        )

        return Response({
            'success': True,
            'source_hopital': {'id': source_hopital.id, 'nom': source_hopital.nom},
            'target_hopital': {'id': target_hopital.id, 'nom': target_hopital.nom},
            'force_center_fields': force,
            'diff': diff,
            'created': created,
            'updated': updated,
        }, status=200)

    @action(detail=True, methods=['post'])
    def copier_referentiel_niveau(self, request, pk=None):
        """Copie la configuration depuis l'hopital de reference du meme niveau."""
        if not request.user.is_superuser:
            raise PermissionDenied("Seul un superuser peut copier le referentiel par niveau.")

        target_hopital = self.get_object()
        force = str(request.data.get('force', 'false')).strip().lower() in {'1', 'true', 'yes', 'oui', 'on'}
        if force and str(request.data.get('confirmation_code') or '').strip().upper() != 'CONFIRMER':
            return Response({'error': 'confirmation_code=CONFIRMER requis quand force=true.'}, status=400)
        source_hopital_id = request.data.get('source_hopital_id')

        if source_hopital_id:
            source_hopital = Hopital.objects.filter(id=source_hopital_id).first()
            if not source_hopital:
                return Response({'error': 'Hôpital source introuvable.'}, status=404)
            if source_hopital.niveau != target_hopital.niveau:
                return Response({'error': 'Le niveau de la source doit correspondre au niveau de la cible.'}, status=400)
        else:
            source_hopital = Hopital.objects.filter(
                niveau=target_hopital.niveau,
                est_reference_niveau=True,
            ).exclude(id=target_hopital.id).order_by('id').first()

        if not source_hopital:
            return Response(
                {'error': f"Aucun hôpital de référence trouvé pour le niveau {target_hopital.get_niveau_display()}."},
                status=400,
            )

        preview_only = str(request.data.get('preview_only', 'false')).strip().lower() in {'1', 'true', 'yes', 'oui', 'on'}
        diff = self._build_configuration_diff(source_hopital, target_hopital)
        if preview_only:
            return Response({
                'success': True,
                'preview_only': True,
                'niveau': target_hopital.niveau,
                'source_hopital': {'id': source_hopital.id, 'nom': source_hopital.nom},
                'target_hopital': {'id': target_hopital.id, 'nom': target_hopital.nom},
                'force_center_fields': force,
                'diff': diff,
            }, status=200)

        created, updated = self._copy_configuration_between_hospitals(
            source_hopital=source_hopital,
            target_hopital=target_hopital,
            force_center_fields=force,
        )
        _log_audit_event(
            request.user,
            'UPDATE',
            'Hopital',
            target_hopital.id,
            f"Copie referentiel niveau {target_hopital.niveau} depuis {source_hopital.code} vers {target_hopital.code} (force={force}).",
        )

        return Response({
            'success': True,
            'niveau': target_hopital.niveau,
            'source_hopital': {'id': source_hopital.id, 'nom': source_hopital.nom},
            'target_hopital': {'id': target_hopital.id, 'nom': target_hopital.nom},
            'force_center_fields': force,
            'diff': diff,
            'created': created,
            'updated': updated,
        }, status=200)

    def _build_configuration_diff(self, source_hopital, target_hopital):
        src_f = set(Fonction.objects.filter(hopital=source_hopital).values_list('code', flat=True))
        tgt_f = set(Fonction.objects.filter(hopital=target_hopital).values_list('code', flat=True))

        src_r = set(CentreResponsabilite.objects.filter(fonction__hopital=source_hopital).values_list('fonction__code', 'code'))
        tgt_r = set(CentreResponsabilite.objects.filter(fonction__hopital=target_hopital).values_list('fonction__code', 'code'))

        src_c = set(CentreCout.objects.filter(centre_responsabilite__fonction__hopital=source_hopital).values_list('centre_responsabilite__fonction__code', 'centre_responsabilite__code', 'code'))
        tgt_c = set(CentreCout.objects.filter(centre_responsabilite__fonction__hopital=target_hopital).values_list('centre_responsabilite__fonction__code', 'centre_responsabilite__code', 'code'))

        src_cp = set(CompteCharge.objects.filter(hopital=source_hopital).values_list('numero', flat=True))
        tgt_cp = set(CompteCharge.objects.filter(hopital=target_hopital).values_list('numero', flat=True))

        diff = {
            'fonctions': {'a_ajouter': len(src_f - tgt_f), 'deja_presentes': len(src_f & tgt_f)},
            'centres_responsabilite': {'a_ajouter': len(src_r - tgt_r), 'deja_presents': len(src_r & tgt_r)},
            'centres_cout': {'a_ajouter': len(src_c - tgt_c), 'deja_presents': len(src_c & tgt_c)},
            'comptes_charges': {'a_ajouter': len(src_cp - tgt_cp), 'deja_presents': len(src_cp & tgt_cp)},
        }
        diff['resume'] = {
            'total_a_ajouter': (
                diff['fonctions']['a_ajouter']
                + diff['centres_responsabilite']['a_ajouter']
                + diff['centres_cout']['a_ajouter']
                + diff['comptes_charges']['a_ajouter']
            ),
            'message': 'Verification terminee. Utiliser preview_only=false pour appliquer.',
        }
        return diff

    @action(detail=False, methods=['get'])
    def controle_qualite_donnees(self, request):
        if not request.user.is_superuser:
            return Response({'error': 'Acces reserve au superuser.'}, status=403)

        anomalies = []
        for hopital in Hopital.objects.order_by('nom'):
            centres = CentreCout.objects.filter(centre_responsabilite__fonction__hopital=hopital)
            comptes_count = CompteCharge.objects.filter(hopital=hopital).count()

            if not Fonction.objects.filter(hopital=hopital).exists():
                anomalies.append({'hopital_id': hopital.id, 'hopital_nom': hopital.nom, 'type': 'fonctions_absentes', 'message': 'Aucune fonction configuree.'})

            if comptes_count == 0:
                anomalies.append({'hopital_id': hopital.id, 'hopital_nom': hopital.nom, 'type': 'comptes_absents', 'message': 'Aucun compte de charge configure.'})

            invalid_uo = centres.filter(type_centre__in=['NT_UO', 'CT_MT', 'CT_CL'], unite_oeuvre='').count()
            if invalid_uo:
                anomalies.append({'hopital_id': hopital.id, 'hopital_nom': hopital.nom, 'type': 'uo_manquante', 'message': f'{invalid_uo} centre(s) avec unite d\'oeuvre manquante.'})

            invalid_tarif = centres.filter(type_centre__in=['CT_MT', 'CT_CL'], tarif__isnull=True).count()
            if invalid_tarif:
                anomalies.append({'hopital_id': hopital.id, 'hopital_nom': hopital.nom, 'type': 'tarif_manquant', 'message': f'{invalid_tarif} centre(s) tarifaire(s) sans tarif.'})

            has_active_ex = Exercice.objects.filter(hopital=hopital, est_actif=True, est_clos=False).exists()
            if not has_active_ex:
                anomalies.append({'hopital_id': hopital.id, 'hopital_nom': hopital.nom, 'type': 'exercice_actif_manquant', 'message': 'Aucun exercice actif non cloture.'})

        return Response({
            'summary': {
                'nb_hopitaux': Hopital.objects.count(),
                'nb_anomalies': len(anomalies),
            },
            'anomalies': anomalies,
        }, status=200)

    @action(detail=False, methods=['get'])
    def alertes_systeme(self, request):
        """Synthese d'alertes operationnelles pour la supervision superuser."""
        if not request.user.is_superuser:
            return Response({'error': 'Acces reserve au superuser.'}, status=403)

        anomalies = self.controle_qualite_donnees(request).data.get('anomalies', [])
        missing_refs = []
        for niveau_code, niveau_label in Hopital.NIVEAU_CHOICES:
            if not Hopital.objects.filter(niveau=niveau_code, est_reference_niveau=True).exists():
                missing_refs.append({'niveau': niveau_code, 'message': f'Pas de reference definie pour {niveau_label}.'})

        since = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        audits_today = AuditLog.objects.filter(timestamp__gte=since).count()

        return Response({
            'summary': {
                'audits_today': audits_today,
                'anomalies_count': len(anomalies),
                'missing_references_count': len(missing_refs),
            },
            'missing_references': missing_refs,
            'anomalies': anomalies,
        }, status=200)

    def _copy_configuration_between_hospitals(self, source_hopital, target_hopital, force_center_fields=False):
        fonction_map = {}
        resp_map = {}
        created = {
            'fonctions': 0,
            'centres_responsabilite': 0,
            'centres_cout': 0,
            'comptes_charges': 0,
        }
        updated = {
            'fonctions': 0,
            'centres_responsabilite': 0,
            'centres_cout': 0,
            'comptes_charges': 0,
        }

        for fonction in Fonction.objects.filter(hopital=source_hopital).order_by('id'):
            new_fonction, is_created = Fonction.objects.get_or_create(
                hopital=target_hopital,
                code=fonction.code,
                defaults={'libelle': fonction.libelle}
            )
            if not is_created and new_fonction.libelle != fonction.libelle:
                new_fonction.libelle = fonction.libelle
                new_fonction.save(update_fields=['libelle'])
                updated['fonctions'] += 1
            if is_created:
                created['fonctions'] += 1
            fonction_map[fonction.id] = new_fonction

        for resp in CentreResponsabilite.objects.filter(fonction__hopital=source_hopital).select_related('fonction').order_by('id'):
            mapped_fonction = fonction_map.get(resp.fonction_id)
            if not mapped_fonction:
                continue
            new_resp, is_created = CentreResponsabilite.objects.get_or_create(
                fonction=mapped_fonction,
                code=resp.code,
                defaults={'libelle': resp.libelle}
            )
            if not is_created and new_resp.libelle != resp.libelle:
                new_resp.libelle = resp.libelle
                new_resp.save(update_fields=['libelle'])
                updated['centres_responsabilite'] += 1
            if is_created:
                created['centres_responsabilite'] += 1
            resp_map[resp.id] = new_resp

        centres_source = CentreCout.objects.filter(
            centre_responsabilite__fonction__hopital=source_hopital
        ).select_related('centre_responsabilite').order_by('id')

        for centre in centres_source:
            mapped_resp = resp_map.get(centre.centre_responsabilite_id)
            if not mapped_resp:
                continue
            new_centre, is_created = CentreCout.objects.get_or_create(
                centre_responsabilite=mapped_resp,
                code=centre.code,
                defaults={
                    'libelle': centre.libelle,
                    'type_centre': centre.type_centre,
                    'unite_oeuvre': centre.unite_oeuvre,
                    'tarif': centre.tarif,
                    'ordre_cascade': centre.ordre_cascade,
                    'est_actif': centre.est_actif,
                }
            )
            if not is_created:
                fields_to_update = []
                if new_centre.libelle != centre.libelle:
                    new_centre.libelle = centre.libelle
                    fields_to_update.append('libelle')

                if force_center_fields:
                    center_fields = [
                        ('type_centre', centre.type_centre),
                        ('unite_oeuvre', centre.unite_oeuvre),
                        ('tarif', centre.tarif),
                        ('ordre_cascade', centre.ordre_cascade),
                        ('est_actif', centre.est_actif),
                    ]
                    for field_name, expected_value in center_fields:
                        if getattr(new_centre, field_name) != expected_value:
                            setattr(new_centre, field_name, expected_value)
                            fields_to_update.append(field_name)

                if fields_to_update:
                    new_centre.save(update_fields=fields_to_update)
                    updated['centres_cout'] += 1
            if is_created:
                created['centres_cout'] += 1

        for compte in CompteCharge.objects.filter(hopital=source_hopital).order_by('id'):
            new_compte, is_created = CompteCharge.objects.get_or_create(
                hopital=target_hopital,
                numero=compte.numero,
                defaults={'libelle': compte.libelle}
            )
            if not is_created and new_compte.libelle != compte.libelle:
                new_compte.libelle = compte.libelle
                new_compte.save(update_fields=['libelle'])
                updated['comptes_charges'] += 1
            if is_created:
                created['comptes_charges'] += 1

        return created, updated

    @action(detail=False, methods=['get'])
    def tableau_bord_superuser(self, request):
        """Indicateurs globaux et alertes pour le superuser."""
        if not request.user.is_superuser:
            return Response({'error': 'Acces reserve au superuser.'}, status=403)

        hopitaux = list(Hopital.objects.all().order_by('nom'))
        hopitaux_ids = [h.id for h in hopitaux]

        exercices_qs = Exercice.objects.filter(hopital_id__in=hopitaux_ids)
        exercices_actifs_qs = exercices_qs.filter(est_actif=True, est_clos=False)

        exercices_par_hopital = defaultdict(list)
        for ex in exercices_qs.order_by('-annee'):
            exercices_par_hopital[ex.hopital_id].append(ex)

        config_stats = {}
        for hopital in hopitaux:
            config_stats[hopital.id] = {
                'fonctions': Fonction.objects.filter(hopital=hopital).count(),
                'centres_responsabilite': CentreResponsabilite.objects.filter(fonction__hopital=hopital).count(),
                'centres_cout': CentreCout.objects.filter(centre_responsabilite__fonction__hopital=hopital).count(),
                'comptes_charges': CompteCharge.objects.filter(hopital=hopital).count(),
            }

        alertes = []
        for hopital in hopitaux:
            has_exercice_actif = Exercice.objects.filter(hopital=hopital, est_actif=True, est_clos=False).exists()
            if not has_exercice_actif:
                alertes.append({
                    'type': 'exercice_actif_manquant',
                    'hopital_id': hopital.id,
                    'hopital_nom': hopital.nom,
                    'message': f"Aucun exercice actif non cloture pour {hopital.nom}."
                })

            cfg = config_stats[hopital.id]
            if cfg['fonctions'] == 0 or cfg['centres_cout'] == 0 or cfg['comptes_charges'] == 0:
                alertes.append({
                    'type': 'configuration_incomplete',
                    'hopital_id': hopital.id,
                    'hopital_nom': hopital.nom,
                    'message': f"Configuration incomplete pour {hopital.nom} (fonctions/centres/comptes)."
                })

        total_charges = Charge.objects.filter(exercice__hopital_id__in=hopitaux_ids).aggregate(total=Sum('montant'))['total'] or Decimal('0')
        total_produits = Produit.objects.filter(exercice__hopital_id__in=hopitaux_ids).aggregate(total=Sum('montant'))['total'] or Decimal('0')
        total_resultat = ResultatCalcul.objects.filter(exercice__hopital_id__in=hopitaux_ids).aggregate(total=Sum('resultat_analytique'))['total'] or Decimal('0')

        return Response({
            'global': {
                'nb_hopitaux': len(hopitaux),
                'nb_exercices_total': exercices_qs.count(),
                'nb_exercices_actifs': exercices_actifs_qs.count(),
                'nb_exercices_clotures': exercices_qs.filter(est_clos=True).count(),
                'total_charges': float(total_charges),
                'total_produits': float(total_produits),
                'total_resultat': float(total_resultat),
                'nb_alertes': len(alertes),
            },
            'alertes': alertes,
            'hopitaux': [
                {
                    'id': h.id,
                    'nom': h.nom,
                    'code': h.code,
                    'exercices': len(exercices_par_hopital.get(h.id, [])),
                    'exercice_actif': next((ex.annee for ex in exercices_par_hopital.get(h.id, []) if ex.est_actif and not ex.est_clos), None),
                    'config': config_stats[h.id],
                }
                for h in hopitaux
            ]
        }, status=200)

    @action(detail=False, methods=['get'])
    def _parse_interhopitaux_filters(self, request):
        annee_param = (request.query_params.get('annee') or '').strip()
        annee = None
        if annee_param:
            try:
                annee = int(annee_param)
            except ValueError:
                raise ValidationError('Parametre annee invalide.')

        actifs_only_raw = (request.query_params.get('actifs_seulement') or '').strip().lower()
        actifs_seulement = actifs_only_raw in ['1', 'true', 'oui', 'yes']

        hopitaux_ids = []
        raw_hopitaux = request.query_params.getlist('hopitaux')
        if not raw_hopitaux:
            raw_single = (request.query_params.get('hopitaux') or '').strip()
            if raw_single:
                raw_hopitaux = [raw_single]

        for raw in raw_hopitaux:
            for token in str(raw).split(','):
                token = token.strip()
                if not token:
                    continue
                try:
                    hopitaux_ids.append(int(token))
                except ValueError:
                    raise ValidationError('Parametre hopitaux invalide. Utiliser des ids entiers.')

        if hopitaux_ids:
            hopitaux_ids = sorted(set(hopitaux_ids))

        niveaux = []
        raw_niveaux = request.query_params.getlist('niveaux')
        if not raw_niveaux:
            raw_single_niveaux = (request.query_params.get('niveaux') or '').strip()
            if raw_single_niveaux:
                raw_niveaux = [raw_single_niveaux]

        valid_niveaux = {choice[0] for choice in Hopital.NIVEAU_CHOICES}
        for raw in raw_niveaux:
            for token in str(raw).split(','):
                token = token.strip().upper()
                if not token:
                    continue
                if token not in valid_niveaux:
                    raise ValidationError("Parametre niveaux invalide. Utiliser N1, N2 ou N3.")
                niveaux.append(token)

        if niveaux:
            niveaux = sorted(set(niveaux))

        return {
            'annee': annee,
            'actifs_seulement': actifs_seulement,
            'hopitaux_ids': hopitaux_ids,
            'niveaux': niveaux,
        }

    def _collect_interhopitaux_data(self, request):
        if not request.user.is_superuser:
            raise PermissionDenied('Acces reserve au superuser.')

        filters = self._parse_interhopitaux_filters(request)
        annee = filters['annee']
        actifs_seulement = filters['actifs_seulement']
        hopitaux_ids = filters['hopitaux_ids']
        niveaux = filters['niveaux']

        hopitaux_qs = Hopital.objects.all().order_by('nom')
        if hopitaux_ids:
            hopitaux_qs = hopitaux_qs.filter(id__in=hopitaux_ids)
        if niveaux:
            hopitaux_qs = hopitaux_qs.filter(niveau__in=niveaux)

        rows = []
        for hopital in hopitaux_qs:
            exercices = Exercice.objects.filter(hopital=hopital)
            if annee is not None:
                exercices = exercices.filter(annee=annee)
            if actifs_seulement:
                exercices = exercices.filter(est_actif=True, est_clos=False)

            if annee is None:
                # Sans annee, privilegier l'exercice actif puis le plus recent.
                exercice_cible = exercices.filter(est_actif=True, est_clos=False).order_by('-annee').first() or exercices.order_by('-annee').first()
                exercices_ids = [exercice_cible.id] if exercice_cible else []
                annee_utilisee = exercice_cible.annee if exercice_cible else None
                exercice_actif = bool(exercice_cible and exercice_cible.est_actif and not exercice_cible.est_clos)
            else:
                exercices_list = list(exercices.order_by('-annee'))
                exercices_ids = [ex.id for ex in exercices_list]
                annee_utilisee = annee
                exercice_actif = any(ex.est_actif and not ex.est_clos for ex in exercices_list)

            if not exercices_ids:
                if actifs_seulement:
                    continue
                rows.append({
                    'hopital_id': hopital.id,
                    'hopital_nom': hopital.nom,
                    'hopital_code': hopital.code,
                    'niveau': hopital.niveau,
                    'niveau_label': hopital.get_niveau_display(),
                    'annee': annee_utilisee,
                    'exercice_actif': False,
                    'charges': 0.0,
                    'produits': 0.0,
                    'resultat': 0.0,
                    'marge': 0.0,
                    'nb_centres': 0,
                })
                continue

            charges = Charge.objects.filter(exercice_id__in=exercices_ids).aggregate(total=Sum('montant'))['total'] or Decimal('0')
            produits = Produit.objects.filter(exercice_id__in=exercices_ids).aggregate(total=Sum('montant'))['total'] or Decimal('0')
            resultat = ResultatCalcul.objects.filter(exercice_id__in=exercices_ids).aggregate(total=Sum('resultat_analytique'))['total'] or Decimal('0')
            nb_centres = ResultatCalcul.objects.filter(exercice_id__in=exercices_ids).values('centre_cout_id').distinct().count()

            charges_float = float(charges)
            produits_float = float(produits)
            resultat_float = float(resultat)
            marge = (resultat_float / produits_float * 100) if produits_float else 0.0

            rows.append({
                'hopital_id': hopital.id,
                'hopital_nom': hopital.nom,
                'hopital_code': hopital.code,
                'niveau': hopital.niveau,
                'niveau_label': hopital.get_niveau_display(),
                'annee': annee_utilisee,
                'exercice_actif': exercice_actif,
                'charges': round(charges_float, 2),
                'produits': round(produits_float, 2),
                'resultat': round(resultat_float, 2),
                'marge': round(marge, 2),
                'nb_centres': nb_centres,
            })

        rows_resultat_desc = sorted(rows, key=lambda item: item['resultat'], reverse=True)
        rows_resultat_asc = sorted(rows, key=lambda item: item['resultat'])

        tot_charges = round(sum(item['charges'] for item in rows), 2)
        tot_produits = round(sum(item['produits'] for item in rows), 2)
        tot_resultat = round(sum(item['resultat'] for item in rows), 2)

        nb_hopitaux = len(rows)
        moyenne_resultat = round((tot_resultat / nb_hopitaux), 2) if nb_hopitaux else 0.0
        meilleur = rows_resultat_desc[0] if rows_resultat_desc else None
        pire = rows_resultat_asc[0] if rows_resultat_asc else None

        alertes = []
        for row in rows:
            if row['produits'] <= 0 and row['charges'] > 0:
                alertes.append({
                    'type': 'produits_nuls',
                    'hopital_id': row['hopital_id'],
                    'message': f"{row['hopital_nom']}: charges sans produits sur le perimetre selectionne.",
                })
            if row['marge'] < 0:
                alertes.append({
                    'type': 'marge_negative',
                    'hopital_id': row['hopital_id'],
                    'message': f"{row['hopital_nom']}: marge negative ({row['marge']}%).",
                })
            if row['nb_centres'] == 0:
                alertes.append({
                    'type': 'aucun_centre_calcule',
                    'hopital_id': row['hopital_id'],
                    'message': f"{row['hopital_nom']}: aucun centre calcule.",
                })

        kpis_par_niveau = {}
        for niveau_code, niveau_libelle in Hopital.NIVEAU_CHOICES:
            rows_niveau = [r for r in rows if r.get('niveau') == niveau_code]
            if not rows_niveau:
                continue
            tot_res = round(sum(item['resultat'] for item in rows_niveau), 2)
            moyenne_res = round(tot_res / len(rows_niveau), 2) if rows_niveau else 0.0
            meilleur_n = max(rows_niveau, key=lambda item: item['resultat'])
            pire_n = min(rows_niveau, key=lambda item: item['resultat'])
            kpis_par_niveau[niveau_code] = {
                'niveau_label': niveau_libelle,
                'nb_hopitaux': len(rows_niveau),
                'resultat_total': tot_res,
                'moyenne_resultat': moyenne_res,
                'ecart_resultat': round(meilleur_n['resultat'] - pire_n['resultat'], 2),
                'meilleur_hopital': {
                    'id': meilleur_n['hopital_id'],
                    'nom': meilleur_n['hopital_nom'],
                    'resultat': meilleur_n['resultat'],
                },
                'pire_hopital': {
                    'id': pire_n['hopital_id'],
                    'nom': pire_n['hopital_nom'],
                    'resultat': pire_n['resultat'],
                },
            }

        return {
            'filters': {
                'annee': annee,
                'actifs_seulement': actifs_seulement,
                'hopitaux_ids': hopitaux_ids,
                'niveaux': niveaux,
            },
            'rows': rows,
            'totaux': {
                'charges': tot_charges,
                'produits': tot_produits,
                'resultat': tot_resultat,
            },
            'kpis': {
                'nb_hopitaux': nb_hopitaux,
                'moyenne_resultat': moyenne_resultat,
                'ecart_resultat': round((meilleur['resultat'] - pire['resultat']), 2) if meilleur and pire else 0.0,
                'meilleur_hopital': {
                    'id': meilleur['hopital_id'],
                    'nom': meilleur['hopital_nom'],
                    'resultat': meilleur['resultat'],
                } if meilleur else None,
                'pire_hopital': {
                    'id': pire['hopital_id'],
                    'nom': pire['hopital_nom'],
                    'resultat': pire['resultat'],
                } if pire else None,
                'nb_alertes': len(alertes),
            },
            'kpis_par_niveau': kpis_par_niveau,
            'top_flop': {
                'top': rows_resultat_desc[:3],
                'flop': rows_resultat_asc[:3],
            },
            'alertes': alertes,
        }

    @action(detail=False, methods=['get'])
    def comparaison_interhopitaux(self, request):
        """Comparaison financiere entre hopitaux pour le superuser."""
        try:
            payload = self._collect_interhopitaux_data(request)
        except PermissionDenied as exc:
            return Response({'error': str(exc)}, status=403)
        except ValidationError as exc:
            detail = exc.detail
            if isinstance(detail, (list, tuple)):
                detail = detail[0]
            return Response({'error': detail}, status=400)

        return Response(payload, status=200)

    @action(detail=False, methods=['get'])
    def export_comparaison_interhopitaux(self, request):
        """Exporte la comparaison inter-hopitaux (filtres courants) en Excel."""
        try:
            payload = self._collect_interhopitaux_data(request)
        except PermissionDenied as exc:
            return Response({'error': str(exc)}, status=403)
        except ValidationError as exc:
            detail = exc.detail
            if isinstance(detail, (list, tuple)):
                detail = detail[0]
            return Response({'error': detail}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Inter Hopitaux'
        ws.append([
            'Hopital', 'Code', 'Niveau', 'Annee', 'Exercice Actif', 'Charges',
            'Produits', 'Resultat', 'Marge (%)', 'Centres Calcules'
        ])

        for row in payload['rows']:
            ws.append([
                row['hopital_nom'],
                row['hopital_code'],
                row.get('niveau_label') or row.get('niveau') or '-',
                row['annee'],
                'Oui' if row.get('exercice_actif') else 'Non',
                row['charges'],
                row['produits'],
                row['resultat'],
                row['marge'],
                row['nb_centres'],
            ])

        ws.append([
            'TOTAL', '', '', '', '',
            payload['totaux']['charges'],
            payload['totaux']['produits'],
            payload['totaux']['resultat'],
            '', ''
        ])

        ws_kpi = wb.create_sheet('KPI')
        ws_kpi.append(['Indicateur', 'Valeur'])
        ws_kpi.append(['Hopitaux compares', payload['kpis']['nb_hopitaux']])
        ws_kpi.append(['Moyenne resultat', payload['kpis']['moyenne_resultat']])
        ws_kpi.append(['Ecart max/min', payload['kpis']['ecart_resultat']])
        ws_kpi.append(['Alertes', payload['kpis']['nb_alertes']])

        ws_niveaux = wb.create_sheet('KPI Par Niveau')
        ws_niveaux.append(['Niveau', 'Hopitaux', 'Resultat Total', 'Moyenne Resultat', 'Ecart'])
        for niveau_code in sorted(payload.get('kpis_par_niveau', {}).keys()):
            item = payload['kpis_par_niveau'][niveau_code]
            ws_niveaux.append([
                item.get('niveau_label', niveau_code),
                item.get('nb_hopitaux', 0),
                item.get('resultat_total', 0),
                item.get('moyenne_resultat', 0),
                item.get('ecart_resultat', 0),
            ])

        _decorate_workbook(wb, 'Comparaison Inter Hopitaux', 'Reseau')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        annee_suffix = payload['filters']['annee'] if payload['filters']['annee'] is not None else 'actif_ou_dernier'
        response['Content-Disposition'] = f'attachment; filename="comparaison_interhopitaux_{annee_suffix}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_comparaison_interhopitaux_pdf(self, request):
        """Exporte la comparaison inter-hopitaux (filtres courants) en PDF."""
        try:
            payload = self._collect_interhopitaux_data(request)
        except PermissionDenied as exc:
            return Response({'error': str(exc)}, status=403)
        except ValidationError as exc:
            detail = exc.detail
            if isinstance(detail, (list, tuple)):
                detail = detail[0]
            return Response({'error': detail}, status=400)

        headers = [
            'Hopital', 'Code', 'Niveau', 'Annee', 'Actif',
            'Charges', 'Produits', 'Resultat', 'Marge %', 'Centres'
        ]
        rows = [
            [
                item['hopital_nom'],
                item['hopital_code'] or '-',
                item.get('niveau_label') or item.get('niveau') or '-',
                item['annee'] or '-',
                'Oui' if item.get('exercice_actif') else 'Non',
                f"{item['charges']:,.0f}",
                f"{item['produits']:,.0f}",
                f"{item['resultat']:,.0f}",
                f"{item['marge']:.2f}",
                str(item['nb_centres']),
            ]
            for item in payload['rows']
        ]
        rows.append([
            'TOTAL', '', '', '', '',
            f"{payload['totaux']['charges']:,.0f}",
            f"{payload['totaux']['produits']:,.0f}",
            f"{payload['totaux']['resultat']:,.0f}",
            '',
            ''
        ])

        annee_suffix = payload['filters']['annee'] if payload['filters']['annee'] is not None else 'actif_ou_dernier'
        return _build_pdf_response(
            filename=f'comparaison_interhopitaux_{annee_suffix}.pdf',
            title='Comparaison Inter Hopitaux',
            hopital_name='Reseau',
            table_headers=headers,
            table_rows=rows,
            section_title='Vue consolidee superuser',
        )
    @action(detail=False, methods=['get'])
    def comparaison_cru_tarifs_interhopitaux(self, request):
        """Compare les CRU et tarifs des centres tarifaires entre hopitaux (superuser)."""
        if not request.user.is_superuser:
            return Response({'error': 'Acces reserve au superuser.'}, status=403)

        filters = self._parse_interhopitaux_filters(request)
        annee = filters['annee']
        actifs_seulement = filters['actifs_seulement']
        hopitaux_ids = filters['hopitaux_ids']

        hopitaux_qs = Hopital.objects.all().order_by('nom')
        if hopitaux_ids:
            hopitaux_qs = hopitaux_qs.filter(id__in=hopitaux_ids)

        rows = []
        columns = []
        for hopital in hopitaux_qs:
            exercices = Exercice.objects.filter(hopital=hopital)
            if annee is not None:
                exercices = exercices.filter(annee=annee)
            if actifs_seulement:
                exercices = exercices.filter(est_actif=True, est_clos=False)

            annee_colonne = annee
            if annee is None:
                exercice_cible = (
                    exercices.filter(est_actif=True, est_clos=False).order_by('-annee').first()
                    or exercices.order_by('-annee').first()
                )
                exercices_ids = [exercice_cible.id] if exercice_cible else []
                if exercice_cible:
                    annee_colonne = exercice_cible.annee
            else:
                exercices_ids = list(exercices.values_list('id', flat=True))

            columns.append({
                'hopital_id': hopital.id,
                'hopital_nom': hopital.nom,
                'hopital_code': hopital.code or '',
                'annee': annee_colonne,
            })

            if not exercices_ids:
                continue

            resultats = (
                ResultatCalcul.objects
                .filter(exercice_id__in=exercices_ids, centre_cout__type_centre__in=['CT_MT', 'CT_CL'])
                .select_related('centre_cout', 'exercice')
                .order_by('exercice__annee', 'centre_cout__code')
            )

            for r in resultats:
                cru = float(r.cout_revient_unitaire) if r.cout_revient_unitaire is not None else None
                tarif = float(r.centre_cout.tarif) if r.centre_cout.tarif is not None else None
                ecart = round(tarif - cru, 2) if cru is not None and tarif is not None else None
                rows.append({
                    'hopital_id': hopital.id,
                    'hopital_nom': hopital.nom,
                    'hopital_code': hopital.code or '',
                    'annee': r.exercice.annee,
                    'centre_code': r.centre_cout.code,
                    'centre_libelle': r.centre_cout.libelle,
                    'type_centre': r.centre_cout.type_centre,
                    'unite_oeuvre': r.centre_cout.unite_oeuvre or '-',
                    'volume_activite': float(r.volume_activite) if r.volume_activite is not None else None,
                    'charges_totales': float(r.charges_totales),
                    'cru': cru,
                    'tarif': tarif,
                    'ecart': ecart,
                })

        return Response({'rows': rows, 'columns': columns, 'filters': filters}, status=200)
